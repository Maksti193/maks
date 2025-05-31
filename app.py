import os
import json
import logging
import datetime
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from openpyxl import Workbook, load_workbook

# Настройка логирования
logging.basicConfig(level=logging.DEBUG)

app = Flask(__name__)
app.secret_key = "your_secret_key"

# Пути к файлам данных
USERS_FILE = "users.json"
TASKS_FILE = "tasks.json"

def load_data(file_path):
    if not os.path.exists(file_path):
        return {} if "users" in file_path else []
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, Exception) as e:
        logging.error(f"Ошибка чтения файла: {str(e)}")
        return {} if "users" in file_path else []

def save_data(data, file_path):
    try:
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception as e:
        logging.error(f"Ошибка сохранения: {str(e)}")

users = load_data(USERS_FILE)
tasks = load_data(TASKS_FILE)

@app.route("/")
def home():
    return render_template("home.html")

@app.route("/auth", methods=["GET", "POST"])
def auth():
    if request.method == "POST":
        login = request.form["login"]
        password = request.form["password"]
        user = users.get(login)
        if user and user["password"] == password:
            session["user"] = login
            session["user_type"] = user["type"]
            return redirect(url_for("task_selection"))
        flash("Неверные учетные данные")
    return render_template("auth.html")

@app.route("/registration", methods=["GET", "POST"])
def registration():
    if request.method == "POST":
        login = request.form["login"]
        password = request.form["password"]
        user_type = int(request.form["user_type"])
        if login not in users:
            users[login] = {"password": password, "type": user_type}
            save_data(users, USERS_FILE)
            flash("Регистрация успешна!")
            return redirect(url_for("auth"))
        flash("Пользователь уже существует")
    return render_template("registration.html")

@app.route("/task_selection")
def task_selection():
    if "user" not in session:
        return redirect(url_for("auth"))
    tasks_stats = {"Титаренко": 1, "Общая сумма штрафов": 0}
    return render_template("task_selection.html", tasks_stats=tasks_stats)

@app.route("/view_all_tasks")
def view_all_tasks():
    if "user" not in session:
        return redirect(url_for("auth"))
    
    def safe_date_parse(task):
        date_str = task["fields"].get("Дата вступления постановления в законную силу", "").strip()
        if not date_str:
            return datetime.datetime.min
        try:
            return datetime.datetime.strptime(date_str, "%d.%m.%Y")
        except ValueError:
            return datetime.datetime.min

    sorted_tasks = sorted(tasks, key=safe_date_parse, reverse=True)
    
    tasks_by_year = {}
    for task in sorted_tasks:
        date_str = task["fields"].get("Дата вступления постановления в законную силу", "").strip()
        try:
            year = datetime.datetime.strptime(date_str, "%d.%m.%Y").year
        except ValueError:
            year = "Без даты"
        
        tasks_by_year.setdefault(year, []).append(task)
    
    sorted_years = sorted(
        [y for y in tasks_by_year if isinstance(y, int)], 
        reverse=True
    )
    if "Без даты" in tasks_by_year:
        sorted_years.append("Без даты")
    
    return render_template(
        "view_all_tasks.html", 
        tasks_by_year=tasks_by_year, 
        sorted_years=sorted_years
    )

@app.route("/article_selection")
def article_selection():
    if "user" not in session:
        return redirect(url_for("auth"))
    return render_template("article_selection.html")

@app.route("/action_selection/<article>")
def action_selection(article):
    if "user" not in session:
        return redirect(url_for("auth"))
    return render_template("action_selection.html", article=article)

@app.route("/create_task/<article>", methods=["GET", "POST"])
def create_task(article):
    if "user" not in session:
        return redirect(url_for("auth"))
    
    if not article:
        flash("Не указан номер статьи!")
        return redirect(url_for("task_selection"))
    
    if request.method == "POST":
        new_task = {
            "id": len(tasks) + 1,
            "article": article,
            "fields": {
                "Номер дела": request.form.get("Номер дела", ""),
                "Краткое наименование": request.form.get("Краткое наименование", ""),
                "Адрес респондента": request.form.get("Адрес респондента", ""),
                "Тип респондента (1 - ЮЛ, 2 - ИП, 3 - ДЛ)": request.form.get("Тип респондента", ""),
                "ИНН": request.form.get("ИНН", ""),
                "ОКПО": request.form.get("ОКПО", ""),
                "Должность руководителя": request.form.get("Должность руководителя", ""),
                "ФИО руководителя": request.form.get("ФИО руководителя", ""),
                "Контактные данные респондента": request.form.get("Контактные данные респондента", ""),
                "Форма стат. отчетности": request.form.get("Форма стат. отчетности", ""),
                "Период представления отчета": request.form.get("Период представления отчета", ""),
                "Сведения о представлении отчета": request.form.get("Сведения о представлении отчета", ""),
                "Дата совершения правонарушения": request.form.get("Дата совершения правонарушения", ""),
                "Должность ответственного за предоставление отчетности": request.form.get("Должность ответственного за предоставление отчетности", ""),
                "ФИО ответственного": request.form.get("ФИО ответственного", ""),
                "Паспортные данные": request.form.get("Паспортные данные", ""),
                "Дата и место рождения": request.form.get("Дата и место рождения", ""),
                "Место жительства": request.form.get("Место жительства", ""),
                "Дата составления протокола об АП, Дата направления определений": request.form.get("Дата составления протокола об АП, Дата направления определений", ""),
                "Сведения о явке на протокол": request.form.get("Сведения о явке на протокол", ""),
                "Отягчающие обстоятельства": request.form.get("Отягчающие обстоятельства", ""),
                "ДАТА РАССМОТРЕНИЕ ДЕЛА": request.form.get("ДАТА РАССМОТРЕНИЕ ДЕЛА", ""),
                "ОТЛОЖЕНИЕ РАССМОТРЕНИЯ (ДАТА)": request.form.get("ОТЛОЖЕНИЕ РАССМОТРЕНИЯ (ДАТА)", ""),
                "Наказание": request.form.get("Наказание", ""),
                "Сумма штрафа": request.form.get("Сумма штрафа", ""),
                "УИН платежа": request.form.get("УИН платежа", ""),
                "Представление об устранении причин": request.form.get("Представление об устранении причин", ""),
                "Дата вступления постановления в законную силу": request.form.get("Дата вступления постановления в законную силу", ""),
                "Сведения о получении ПОСТАНОВЛЕНИЯ": request.form.get("Сведения о получении ПОСТАНОВЛЕНИЯ", ""),
                "Сведения об ответе на представление": request.form.get("Сведения об ответе на представление", ""),
                "Сведения об оплате": request.form.get("Сведения об оплате", ""),
                "Сумма оплаченных штрафов": request.form.get("Сумма оплаченных штрафов", ""),
                "Сведения об обжаловании": request.form.get("Сведения об обжаловании", ""),
                "Дата направления приставам на взыскание": request.form.get("Дата направления приставам на взыскание", "")
            }
        }
        tasks.append(new_task)
        save_data(tasks, TASKS_FILE)
        flash("Запись успешно создана!")
        return redirect(url_for("action_selection", article=article))
    
    return render_template("create_task.html", article=article)

@app.route("/import_from_excel", methods=["GET", "POST"])
def import_from_excel():
    if "user" not in session:
        return redirect(url_for("auth"))
    
    if request.method == "POST":
        if 'file' not in request.files:
            flash('Файл не выбран')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('Файл не выбран')
            return redirect(request.url)
        
        if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            try:
                wb = load_workbook(file, read_only=True)
                ws = wb.active
                
                headers = [cell.value for cell in ws[1]]
                
                field_mapping = {
                    "Номер дела": "Номер дела",
                    "Краткое наименование": "Краткое наименование",
                    "Адрес респондента": "Адрес респондента",
                    "Тип респондента": "Тип респондента (1 - ЮЛ, 2 - ИП, 3 - ДЛ)",
                    "ИНН": "ИНН",
                    "ОКПО": "ОКПО",
                    "Должность руководителя": "Должность руководителя",
                    "ФИО руководителя": "ФИО руководителя",
                    "Контактные данные респондента": "Контактные данные респондента",
                    "Форма стат. отчетности": "Форма стат. отчетности",
                    "Период представления отчета": "Период представления отчета",
                    "Сведения о представлении отчета": "Сведения о представлении отчета",
                    "Дата совершения правонарушения": "Дата совершения правонарушения",
                    "Должность ответственного за предоставление отчетности": "Должность ответственного за предоставление отчетности",
                    "ФИО ответственного": "ФИО ответственного",
                    "Паспортные данные": "Паспортные данные",
                    "Дата и место рождения": "Дата и место рождения",
                    "Место жительства": "Место жительства",
                    "Дата составления протокола об АП, Дата направления определений": "Дата составления протокола об АП, Дата направления определений",
                    "Сведения о явке на протокол": "Сведения о явке на протокол",
                    "Отягчающие обстоятельства": "Отягчающие обстоятельства",
                    "ДАТА РАССМОТРЕНИЕ ДЕЛА": "ДАТА РАССМОТРЕНИЕ ДЕЛА",
                    "ОТЛОЖЕНИЕ РАССМОТРЕНИЯ (ДАТА)": "ОТЛОЖЕНИЕ РАССМОТРЕНИЯ (ДАТА)",
                    "Наказание": "Наказание",
                    "Сумма штрафа": "Сумма штрафа",
                    "УИН платежа": "УИН платежа",
                    "Представление об устранении причин": "Представление об устранении причин",
                    "Дата вступления постановления в законную силу": "Дата вступления постановления в законную силу",
                    "Сведения о получении ПОСТАНОВЛЕНИЯ": "Сведения о получении ПОСТАНОВЛЕНИЯ",
                    "Сведения об ответе на представление": "Сведения об ответе на представление",
                    "Сведения об оплате": "Сведения об оплате",
                    "Сумма оплаченных штрафов": "Сумма оплаченных штрафов",
                    "Сведения об обжаловании": "Сведения об обжаловании",
                    "Дата направления приставам на взыскание": "Дата направления приставам на взыскание"
                }
                
                imported_count = 0
                existing_case_numbers = {t["fields"].get("Номер дела") for t in tasks}
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                        
                    new_task = {
                        "id": len(tasks) + 1,
                        "article": request.form.get("article", "13.19"),
                        "fields": {}
                    }
                    
                    for header, value in zip(headers, row):
                        if header in field_mapping and value is not None:
                            new_task["fields"][field_mapping[header]] = str(value).strip()
                    
                    # Автоматически устанавливаем "Штраф" если в поле "Сумма штрафа" есть числовое значение
                    if "Сумма штрафа" in new_task["fields"]:
                        try:
                            penalty_amount = new_task["fields"]["Сумма штрафа"].replace(",", ".")
                            float(penalty_amount)
                            new_task["fields"]["Наказание"] = "Штраф"
                        except ValueError:
                            pass
                    
                    case_number = new_task["fields"].get("Номер дела")
                    if case_number and case_number not in existing_case_numbers:
                        tasks.append(new_task)
                        existing_case_numbers.add(case_number)
                        imported_count += 1
                
                save_data(tasks, TASKS_FILE)
                flash(f"Успешно импортировано {imported_count} записей!")
                return redirect(url_for("task_selection"))
            
            except Exception as e:
                flash(f"Ошибка при импорте файла: {str(e)}")
                return redirect(request.url)
        else:
            flash('Неподдерживаемый формат файла. Используйте .xlsx или .xls')
            return redirect(request.url)
    
    return render_template("import_from_excel.html")

@app.route("/search_records/<article>", methods=["GET"])
def search_records(article):
    query = request.args.get("query", "").strip().lower()
    if len(query) < 3:
        return jsonify({"error": "Слишком короткий запрос"}), 400

    results = [
        {
            "id": record["id"],
            "номер_дела": record["fields"].get("Номер дела", ""),
            "краткое_наименование": record["fields"].get("Краткое наименование", ""),
            "дата_рассмотрения": record["fields"].get("ДАТА РАССМОТРЕНИЕ ДЕЛА", "")
        }
        for record in tasks
        if record.get("article") == article
        and any(query in str(value).lower() for value in record.values())
    ]

    results.sort(key=lambda x: x["дата_рассмотрения"] or "")
    return jsonify(results)

@app.route("/view_task/<int:task_id>")
def view_task(task_id):
    if "user" not in session:
        return redirect(url_for("auth"))
    task = next((t for t in tasks if t["id"] == task_id), None)
    if not task:
        flash("Запись не найдена!")
        return redirect(url_for("task_selection"))
    return render_template("view_task.html", task=task)

@app.route("/edit_task/<int:task_id>", methods=["GET", "POST"])
def edit_task(task_id):
    if "user" not in session:
        return redirect(url_for("auth"))
    task = next((t for t in tasks if t["id"] == task_id), None)
    if not task:
        flash("Запись не найдена!")
        return redirect(url_for("task_selection"))
    
    if request.method == "POST":
        task["fields"] = {
            "Номер дела": request.form.get("Номер дела", ""),
            "Краткое наименование": request.form.get("Краткое наименование", ""),
            "Адрес респондента": request.form.get("Адрес респондента", ""),
            "Тип респондента (1 - ЮЛ, 2 - ИП, 3 - ДЛ)": request.form.get("Тип респондента", ""),
            "ИНН": request.form.get("ИНН", ""),
            "ОКПО": request.form.get("ОКПО", ""),
            "Должность руководителя": request.form.get("Должность руководителя", ""),
            "ФИО руководителя": request.form.get("ФИО руководителя", ""),
            "Контактные данные респондента": request.form.get("Контактные данные респондента", ""),
            "Форма стат. отчетности": request.form.get("Форма стат. отчетности", ""),
            "Период представления отчета": request.form.get("Период представления отчета", ""),
            "Сведения о представлении отчета": request.form.get("Сведения о представлении отчета", ""),
            "Дата совершения правонарушения": request.form.get("Дата совершения правонарушения", ""),
            "Должность ответственного за предоставление отчетности": request.form.get("Должность ответственного за предоставление отчетности", ""),
            "ФИО ответственного": request.form.get("ФИО ответственного", ""),
            "Паспортные данные": request.form.get("Паспортные данные", ""),
            "Дата и место рождения": request.form.get("Дата и место рождения", ""),
            "Место жительства": request.form.get("Место жительства", ""),
            "Дата составления протокола об АП, Дата направления определений": request.form.get("Дата составления протокола об АП, Дата направления определений", ""),
            "Сведения о явке на протокол": request.form.get("Сведения о явке на протокол", ""),
            "Отягчающие обстоятельства": request.form.get("Отягчающие обстоятельства", ""),
            "ДАТА РАССМОТРЕНИЕ ДЕЛА": request.form.get("ДАТА РАССМОТРЕНИЕ ДЕЛА", ""),
            "ОТЛОЖЕНИЕ РАССМОТРЕНИЯ (ДАТА)": request.form.get("ОТЛОЖЕНИЕ РАССМОТРЕНИЯ (ДАТА)", ""),
            "Наказание": request.form.get("Наказание", ""),
            "Сумма штрафа": request.form.get("Сумма штрафа", ""),
            "УИН платежа": request.form.get("УИН платежа", ""),
            "Представление об устранении причин": request.form.get("Представление об устранении причин", ""),
            "Дата вступления постановления в законную силу": request.form.get("Дата вступления постановления в законную силу", ""),
            "Сведения о получении ПОСТАНОВЛЕНИЯ": request.form.get("Сведения о получении ПОСТАНОВЛЕНИЯ", ""),
            "Сведения об ответе на представление": request.form.get("Сведения об ответе на представление", ""),
            "Сведения об оплате": request.form.get("Сведения об оплате", ""),
            "Сумма оплаченных штрафов": request.form.get("Сумма оплаченных штрафов", ""),
            "Сведения об обжаловании": request.form.get("Сведения об обжаловании", ""),
            "Дата направления приставам на взыскание": request.form.get("Дата направления приставам на взыскание", "")
        }
        save_data(tasks, TASKS_FILE)
        flash("Запись успешно обновлена!")
        return redirect(url_for("action_selection", article=task["article"]))
    
    return render_template("edit_task.html", task=task)

@app.route("/delete_task/<int:task_id>", methods=["POST"])
def delete_task(task_id):
    if "user" not in session:
        return redirect(url_for("auth"))
    
    global tasks
    tasks = load_data(TASKS_FILE)
    
    task = next((t for t in tasks if t["id"] == task_id), None)
    if not task:
        flash("Запись не найдена!")
        return redirect(url_for("task_selection"))
    
    tasks = [t for t in tasks if t["id"] != task_id]
    save_data(tasks, TASKS_FILE)
    
    flash("Запись успешно удалена!")
    return redirect(url_for("action_selection", article=task["article"]))

@app.route("/export_to_excel")
def export_to_excel():
    if "user" not in session:
        return redirect(url_for("auth"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Сведения о делах"

    headers = [
        "Номер дела",
        "Краткое наименование",
        "Адрес респондента",
        "Тип респондента (1 - ЮЛ, 2 - ИП, 3 - ДЛ)",
        "ИНН",
        "ОКПО",
        "Должность руководителя",
        "ФИО руководителя",
        "Контактные данные респондента",
        "Форма стат. отчетности",
        "Период представления отчета",
        "Сведения о представлении отчета",
        "Дата совершения правонарушения",
        "Должность ответственного за предоставление отчетности",
        "ФИО ответственного",
        "Паспортные данные",
        "Дата и место рождения",
        "Место жительства",
        "Дата составления протокола об АП, Дата направления определений",
        "Сведения о явке на протокол",
        "Отягчающие обстоятельства",
        "ДАТА РАССМОТРЕНИЕ ДЕЛА",
        "ОТЛОЖЕНИЕ РАССМОТРЕНИЯ (ДАТА)",
        "Наказание",
        "Сумма штрафа",
        "УИН платежа",
        "Представление об устранении причин",
        "Дата вступления постановления в законную силу",
        "Сведения о получении ПОСТАНОВЛЕНИЯ",
        "Сведения об ответе на представление",
        "Сведения об оплате",
        "Сумма оплаченных штрафов",
        "Сведения об обжаловании",
        "Дата направления приставам на взыскание"
    ]

    ws.append(headers)
    for task in tasks:
        row = [task["fields"].get(header, "") for header in headers]
        ws.append(row)

    excel_file = "сведения_о_делах.xlsx"
    wb.save(excel_file)

    return send_file(
        excel_file,
        as_attachment=True,
        download_name="сведения_о_делах.xlsx"
    )

@app.route("/repetition_check", methods=["GET", "POST"])
def repetition_check():
    if request.method == "POST":
        okpo = request.form.get("okpo", "").strip()
        violation_date_str = request.form.get("violation_date", "").strip()
        
        try:
            violation_date = datetime.datetime.strptime(violation_date_str, "%d.%m.%Y").date()
        except ValueError:
            flash("Некорректный формат даты! Используйте ДД.ММ.ГГГГ", "error")
            return render_template("repetition_check.html")
        
        matching_tasks = [
            task for task in tasks 
            if (task.get("article") == "13.19" and 
                str(task["fields"].get("ОКПО", "")).strip() == okpo and
                task["fields"].get("Наказание") != "Прекращение")
        ]
        
        results = []
        is_repeat = False
        
        for task in matching_tasks:
            date_str = task["fields"].get("Дата вступления постановления в законную силу", "")
            try:
                effective_date = datetime.datetime.strptime(date_str, "%d.%m.%Y").date()
                delta = (violation_date - effective_date).days
                if 0 <= delta <= 365:
                    is_repeat = True
                    results.append({
                        "case_number": task["fields"].get("Номер дела", "Не указан"),
                        "effective_date": date_str,
                        "punishment": task["fields"].get("Наказание", "Не указано")
                    })
            except (ValueError, TypeError):
                continue
        
        return render_template(
            "repetition_check.html",
            okpo=okpo,
            violation_date=violation_date_str,
            is_repeat=is_repeat,
            results=results
        )
    
    return render_template("repetition_check.html")

@app.route("/logout")
def logout():
    session.pop("user", None)
    session.pop("user_type", None)
    return redirect(url_for("home"))

if __name__ == "__main__":
    app.run(debug=True, port=8000)